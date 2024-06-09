import openpyxl, time
t0 = time.time()
wb = openpyxl.load_workbook(['test.xlsx','恐龙.xlsx'][1])
t1 = time.time()
print('successfully opened')
print()
worksheet = wb.worksheets[0]
mr = worksheet.max_row
rows = worksheet.iter_rows(1,mr)
table = []
row = 0
for i in rows:
    row += 1
    print(f'{row}/{mr}')
    r = []
    for c in i:
        r.append(c.value)
    table.append(r)
print()
t2 = time.time()
print(t1-t0, t2-t1, 'Total:', t2-t0)